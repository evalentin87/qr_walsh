import os, re, io, json, shutil
from urllib.parse import quote_plus
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, abort, Response
import qrcode
from PIL import Image
import openpyxl
import requests
import os

# Si NO hay variables de entorno, usará las carpetas locales (funciona igual en Windows)
BASE_DIR   = os.environ.get("QR_BASE_DIR", os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.environ.get("QR_DATA_DIR",   os.path.join(BASE_DIR, "data"))
PHOTOS_DIR = os.environ.get("QR_PHOTOS_DIR", os.path.join(BASE_DIR, "photos"))
QRS_DIR    = os.environ.get("QR_QRS_DIR",    os.path.join(BASE_DIR, "qrs"))
VCF_DIR    = os.environ.get("QR_VCF_DIR",    os.path.join(BASE_DIR, "vcards"))
STATIC_DIR = os.path.join(BASE_DIR, "static")

for p in [DATA_DIR, PHOTOS_DIR, QRS_DIR, VCF_DIR]:
    os.makedirs(p, exist_ok=True)

app = Flask(__name__)

def slugify(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s.strip())
    return re.sub(r"_+", "_", s).strip("_")

def person_path(slug): return os.path.join(DATA_DIR, f"{slug}.json")
def photo_path(slug): return os.path.join(PHOTOS_DIR, f"{slug}.jpg")
def qr_path(slug):    return os.path.join(QRS_DIR, f"{slug}.png")
def vcf_path(slug):   return os.path.join(VCF_DIR, f"{slug}.vcf")

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/form")
def form_view():
    return render_template("form.html")

def save_json(slug, data):
    with open(person_path(slug), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_json(slug):
    with open(person_path(slug), "r", encoding="utf-8") as f:
        return json.load(f)

def store_photo(slug, foto_url_or_path, file_storage):
    # Uploaded file
    if file_storage and getattr(file_storage, "filename", ""):
        file_storage.save(photo_path(slug))
        return True
    # Local absolute
    if foto_url_or_path and os.path.isabs(foto_url_or_path) and os.path.exists(foto_url_or_path):
        shutil.copy(foto_url_or_path, photo_path(slug))
        return True
    # Remote URL
    if foto_url_or_path and foto_url_or_path.lower().startswith(("http://", "https://")):
        try:
            import requests
            r = requests.get(foto_url_or_path, timeout=10)
            if r.status_code == 200:
                with open(photo_path(slug), "wb") as f:
                    f.write(r.content)
                return True
        except Exception:
            pass
    return False

# ---------------------------
# Helpers para vCard (NUEVO)
# ---------------------------
def _clean(s):
    return (s or "").strip()

def _phone_e164(s):
    s = str(s or "")
    keep_plus = s.strip().startswith("+")
    digits = re.sub(r"\D+", "", s)
    return ("+" if keep_plus else "") + digits

def _build_vcard(p, slug=None, org_name="Walsh Perú"):
    """Devuelve la vCard 3.0 con CRLF obligatorio."""
    nombre = _clean(p.get("nombre"))
    apellido = _clean(p.get("apellido"))
    cargo = _clean(p.get("cargo"))
    area = _clean(p.get("area"))
    correo = _clean(p.get("correo"))
    celular = _phone_e164(p.get("celular"))
    direccion = _clean(p.get("direccion"))
    web = _clean(p.get("web"))

    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"N:{apellido};{nombre};;;",
        f"FN:{nombre} {apellido}",
    ]
    # ORG/TITLE si existen
    if org_name or area:
        lines.append(f"ORG:{org_name};{area}")
    if cargo:
        lines.append(f"TITLE:{cargo}")
    # Teléfono y correo
    if celular:
        lines.append(f"TEL;TYPE=CELL,VOICE:{celular}")
    if correo:
        lines.append(f"EMAIL;TYPE=INTERNET,PREF:{correo}")
    # Dirección y Web
    if direccion:
        lines.append(f"ADR;TYPE=WORK:;;{direccion};;;;")
    if web:
        lines.append(f"URL:{web}")
    # Foto por URL (opcional)
    try:
        if p.get("foto_local") and slug:
            abs_photo = request.url_root.rstrip("/") + url_for("photo", slug=slug)
            lines.append(f"PHOTO;VALUE=URI:{abs_photo}")
        elif p.get("foto"):
            lines.append(f"PHOTO;VALUE=URI:{_clean(p.get('foto'))}")
    except Exception:
        pass

    lines.append("END:VCARD")
    return "\r\n".join(lines) + "\r\n"

def make_vcard(slug, p):
    """Escribe el .vcf en disco (útil si luego quieres empaquetar)."""
    vcf_text = _build_vcard(p, slug)
    with open(vcf_path(slug), "wb") as f:
        f.write(vcf_text.encode("utf-8"))

def make_qr(slug):
    url = url_for("icard", slug=slug, _external=True)
    img = qrcode.make(url)
    img.save(qr_path(slug))

def build_person_from_form(form):
    p = {
        "nombre": form.get("nombre","").strip(),
        "apellido": form.get("apellido","").strip(),
        "cargo": form.get("cargo","").strip(),
        "area": form.get("area","").strip(),
        "correo": form.get("correo","").strip(),
        "celular": form.get("celular","").strip(),
        "direccion": form.get("direccion","").strip(),
        "web": form.get("web","").strip(),
        "foto": form.get("foto","").strip(),
    }
    return p

@app.post("/create_manual")
def create_manual():
    p = build_person_from_form(request.form)
    slug = slugify(f"{p.get('nombre','')}_{p.get('apellido','')}") or "tarjeta"
    # Photo
    store_photo(slug, p.get("foto"), request.files.get("foto_archivo"))
    if os.path.exists(photo_path(slug)):
        p["foto_local"] = True
    save_json(slug, p)
    make_vcard(slug, p)
    make_qr(slug)
    return redirect(url_for("result_view", slug=slug))

@app.post("/create_bulk")
def create_bulk():
    f = request.files.get("xlsx")
    if not f: abort(400, "Falta archivo Excel")
    tmp = os.path.join(DATA_DIR, "_bulk.xlsx")
    f.save(tmp)

    import openpyxl
    wb = openpyxl.load_workbook(tmp)
    ws = wb.active
    headers = [ (c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) ]
    fields = ["nombre","apellido","cargo","area","correo","celular","direccion","web","foto"]
    idx = { k: headers.index(k) if k in headers else None for k in fields }

    for row in ws.iter_rows(min_row=2, values_only=True):
        p = { k: (row[idx[k]] if idx[k] is not None else "") for k in fields }
        p = {k:(v or "").strip() for k,v in p.items()}
        slug = slugify(f"{p.get('nombre','')}_{p.get('apellido','')}") or "tarjeta"
        store_photo(slug, p.get("foto"), None)
        if os.path.exists(photo_path(slug)): p["foto_local"] = True
        save_json(slug, p)
        make_vcard(slug, p)
        make_qr(slug)
    return redirect(url_for("form_view"))

@app.get("/card/<slug>")
def card_view(slug):
    p = load_json(slug)
    tel = f"tel:{p.get('celular','')}"
    mail = f"mailto:{p.get('correo','')}"
    whatsapp = f"https://wa.me/{p.get('celular','').replace('+','').replace(' ','')}"
    maps = f"https://www.google.com/maps/search/?api=1&query={quote_plus(p.get('direccion',''))}"
    qr_url = url_for("download_qr", slug=slug)
    vcf_url = url_for("download_vcf", slug=slug)
    return render_template("card.html", person=p, slug=slug,
                           tel_link=tel, mail_link=mail, whatsapp_link=whatsapp,
                           maps_link=maps, qr_url=qr_url, vcf_url=vcf_url,
                           card_jpg_url="#")

@app.get("/icard/<slug>")
def icard(slug):
    p = load_json(slug)
    return render_template("icard.html", person=p, slug=slug)

@app.get("/result/<slug>")
def result_view(slug):
    p = load_json(slug)
    qr_url = url_for("download_qr", slug=slug)
    vcf_url = url_for("download_vcf", slug=slug)
    return render_template("result.html", person=p, slug=slug,
                           qr_url=qr_url, vcf_url=vcf_url)

@app.get("/download_qr/<slug>")
def download_qr(slug):
    path = qr_path(slug)
    if not os.path.exists(path): abort(404)
    return send_from_directory(QRS_DIR, f"{slug}.png", as_attachment=True)

# ------------ DESCARGA VCF (NUEVO) ------------
@app.get("/download_vcf/<slug>")
def download_vcf(slug):
    p = load_json(slug)
    p["slug"] = slug
    vcf_text = _build_vcard(p, slug)
    filename = f"{_clean(p.get('nombre'))}_{_clean(p.get('apellido'))}.vcf".strip() or f"{slug}.vcf"
    headers = {
        "Content-Type": "text/vcard; charset=utf-8",
        "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
        "X-Content-Type-Options": "nosniff",
    }
    return Response(vcf_text.encode("utf-8"), headers=headers)
# ----------------------------------------------

@app.get("/photo/<slug>")
def photo(slug):
    path = photo_path(slug)
    if not os.path.exists(path): abort(404)
    return send_from_directory(PHOTOS_DIR, f"{slug}.jpg")

@app.get("/logo.png")
def serve_logo():
    return send_from_directory(os.path.join(STATIC_DIR, "img"), "Logo.png")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
